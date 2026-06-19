from openpyxl import load_workbook, Workbook


def read_excel(path):
    wb = load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], sheet_names
    header = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = []
    for row_vals in rows[1:]:
        row_dict = {}
        for i, col_name in enumerate(header):
            val = row_vals[i] if i < len(row_vals) else None
            row_dict[col_name] = str(val) if val is not None else ""
        if any(v.strip() for v in row_dict.values()):
            data_rows.append(row_dict)
    wb.close()
    return data_rows, header, sheet_names


def read_sheet(path, sheet_name):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return [], []
    header = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = []
    for row_vals in rows[1:]:
        row_dict = {}
        for i, col_name in enumerate(header):
            val = row_vals[i] if i < len(row_vals) else None
            row_dict[col_name] = str(val) if val is not None else ""
        if any(v.strip() for v in row_dict.values()):
            data_rows.append(row_dict)
    wb.close()
    return data_rows, header


def write_excel(path, rows, columns):
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
    wb.save(path)
    wb.close()
